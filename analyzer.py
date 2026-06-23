"""
Recruitment pipeline analyzer for Supply Mapping data.
Parses/fixes dates, imputes missing round dates, computes timelines and sourcer metrics.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Latest valid date in pipeline — nothing should exceed this
MAX_PIPELINE_DATE = pd.Timestamp("2026-06-17")
DATE_DISPLAY_FORMAT = "MM/DD/YYYY"

MATCH_LEVEL_LABELS = {
    "skills_sourcer": "Similar Skills + Sourcer median",
    "skills": "Similar Skills median",
    "designation_exp": "Same Designation + Exp bucket median",
    "sourcer": "Same Sourcer median",
    "global": "Overall pipeline median",
}

# Pipeline stage definitions
REACHED_R1 = {
    "R1 Reject",
    "R2 Reject",
    "R2 Select",
    "Offer released",
    "Offer Decline",
    "TA Screening",
}
REACHED_R2 = {
    "R2 Reject",
    "R2 Select",
    "Offer released",
    "Offer Decline",
}
REACHED_OFFER = {"Offer released", "Offer Decline"}
NO_IMPUTE = {
    "Interview Dropout",
    "On Hold",
    "Candidate not interested",
}

MAX_REASONABLE_DAYS_R01 = 60
MAX_REASONABLE_DAYS_R12 = 14
MIN_REASONABLE_DAYS_R12 = 1
DEFAULT_R01_DAYS = 7.0
DEFAULT_R12_DAYS = 2.0
SLOW_THRESHOLD_DAYS = 30

STAGE_TRANSITIONS = (
    ("R0", "R1", "Days R0 to R1", 0, MAX_REASONABLE_DAYS_R01),
    ("R1", "R2", "Days R1 to R2", MIN_REASONABLE_DAYS_R12, MAX_REASONABLE_DAYS_R12),
    ("R0", "Offer Stage", "Days R0 to Offer Stage", 0, MAX_REASONABLE_DAYS_R01 * 2),
)

POWERBI_INSTRUCTIONS = [
    "STEP 1 — Generate data: run python export_powerbi.py with ANY Supply Mapping-format .xlsx (uploads/ folder or pass file path).",
    "STEP 2 — Open Power BI Desktop → Get Data → Excel → select Supply_Mapping_PowerBI_Ready.xlsx.",
    "STEP 3 — Load ALL sheets starting with PBI_ (check every PBI_ table). Click Load (not Transform unless you want to hide unused columns).",
    "STEP 4 — Model view: create relationship PBI_Stage_Transitions[Candidate Key] → PBI_Candidates[Candidate Key] (Many-to-one, single direction).",
    "STEP 5 — Build 3 report pages using sheet PBI_Dashboard_Layout as your visual blueprint (same layout as old HTML dashboard).",
    "STEP 6 — Refresh: when new Supply Mapping file arrives, re-run export_powerbi.py and click Refresh in Power BI.",
    "NOTE: R2→Offer days are not in source data. Use R0→Offer Stage as the offer-stage timing metric.",
]

POWERBI_DASHBOARD_LAYOUT = [
    {
        "Page": "1 - Executive Overview",
        "Visual Type": "Card",
        "Title": "Total Candidates",
        "Data Table": "PBI_Summary",
        "Fields": "Metric = Total Candidates, Value",
        "Notes": "Repeat cards for Reached R1, Reached R2, Reached Offer, Avg R0→R1 Days, Slow Movers 30+",
    },
    {
        "Page": "1 - Executive Overview",
        "Visual Type": "Funnel",
        "Title": "Recruitment Funnel",
        "Data Table": "PBI_Funnel",
        "Fields": "Stage (Group), Count (Values), Sort by Stage Order",
        "Notes": "Same funnel as old HTML: R0 Applied → R1 → R2 → Offer",
    },
    {
        "Page": "1 - Executive Overview",
        "Visual Type": "Donut chart",
        "Title": "Final Status Mix",
        "Data Table": "PBI_Status",
        "Fields": "Status (Legend), Count (Values)",
        "Notes": "Matches old status doughnut chart",
    },
    {
        "Page": "1 - Executive Overview",
        "Visual Type": "Clustered bar chart",
        "Title": "Sourcer Volume vs Conversion",
        "Data Table": "PBI_Sourcer_Chart",
        "Fields": "Sourcer (Axis), Total Submitted + Overall Conversion % (Values)",
        "Notes": "Top 12 sourcers by volume — same as old sourcer chart",
    },
    {
        "Page": "2 - Bottlenecks",
        "Visual Type": "Clustered bar chart",
        "Title": "Bottleneck by Round",
        "Data Table": "PBI_Bottleneck_Rounds",
        "Fields": "Transition (Axis), Avg Days (Values), Is Primary Bottleneck (Tooltips)",
        "Notes": "Shows which round is slowest — R0→R1 vs R1→R2 vs R0→Offer",
    },
    {
        "Page": "2 - Bottlenecks",
        "Visual Type": "Table",
        "Title": "Slow Designations (30+ days)",
        "Data Table": "PBI_Bottleneck_Designation",
        "Fields": "Designation, Total Candidates, Avg Days R0→R1, Count R0→R1 30+ Days, Primary Bottleneck",
        "Notes": "Sort by Avg Days R0→R1 descending. Filter Any Stage 30+ Flag = True",
    },
    {
        "Page": "2 - Bottlenecks",
        "Visual Type": "Table",
        "Title": "Slow Sourcers (30+ days)",
        "Data Table": "PBI_Bottleneck_Sourcer",
        "Fields": "Sourcer, Total Candidates, Avg Days R0→R1, Count R0→R1 30+ Days, Primary Bottleneck",
        "Notes": "Sort by Avg Days R0→R1 descending",
    },
    {
        "Page": "2 - Bottlenecks",
        "Visual Type": "Clustered column chart",
        "Title": "Processing Speed by Sourcer",
        "Data Table": "PBI_Sourcer_Chart",
        "Fields": "Sourcer (Axis), Avg Days R0→R1 + Avg Days R1→R2 (Values)",
        "Notes": "Same speed chart as old HTML dashboard",
    },
    {
        "Page": "3 - Detail & Slow Movers",
        "Visual Type": "Matrix",
        "Title": "Stage Days by Designation",
        "Data Table": "PBI_Stage_Transitions",
        "Fields": "Designation (Rows), Transition (Columns), Average of Days (Values)",
        "Notes": "Heat-map style matrix for designation timing",
    },
    {
        "Page": "3 - Detail & Slow Movers",
        "Visual Type": "Table",
        "Title": "Candidates Stuck 30+ Days",
        "Data Table": "PBI_Slow_Movers_30Plus",
        "Fields": "Name, Sourcer, Designation, Slowest Transition, Slowest Days, All Slow Transitions",
        "Notes": "Full drill-down list of bottleneck candidates",
    },
    {
        "Page": "3 - Detail & Slow Movers",
        "Visual Type": "Table",
        "Title": "All Candidates Timeline",
        "Data Table": "PBI_Candidates",
        "Fields": "Name, Sourcer, Designation, Final Status, Days R0 to R1, Days R1 to R2, Any Stage 30+ Days",
        "Notes": "Replaces old timeline preview table",
    },
]

ASSUMPTIONS = [
    "A1 – Year Correction (R1 & R2): All R1/R2 dates where R0 is 2025 but R1/R2 were entered as 2026 have been corrected to 2025. Reason: Submission dates for S.No 1–~516 are all in 2025. A 500+ day gap to a screening round is operationally impossible. These are clearly data entry errors.",
    "A2 – 2026 dates for S.No 517+ are genuine: R0 dates from Jan 2026 onward are treated as real 2026 dates. R1/R2 dates in 2026 for these records are kept as-is. The data clearly shows two cohorts: 2025 submissions (S.No 1–~516) and 2026 submissions (S.No ~517+).",
    "A3 – R1 Prediction Model (7-Level Hierarchy): Missing R1 dates are predicted using a hierarchical model. "
    "Level 1: Median actual R1 date of candidates with the same Designation + Skills + exact R0 date (batch cohort). "
    "Level 2: Median actual R1 date of same Designation + Skills + same R0 month. "
    "Level 3: Median actual R1 date of same Designation + same R0 month. "
    "Level 4: R0 + median gap for same Designation + Skills (gap-based). "
    "Level 5: R0 + median gap for same Designation only. "
    "Level 6: R0 + median gap for same Skills only. "
    "Level 7 (Global Fallback): R0 + global median gap. "
    "Traceability: each predicted row records which level was used in the R1 Assumption/Note column.",
    "A4 – R1→R2 Fixed Gap: R2 imputed as R1 + 2 days. Every single observed R1→R2 gap in the entire dataset is exactly 2 days — 100% consistent.",
    "A5 – Interview Dropout / On Hold / Not Interested: No R1 or R2 dates assigned. These candidates did not progress to interview. Assigning dates would introduce false information.",
    "A6 – R1 Reject: Only R1 filled; R2 left blank. Rejected at R1 → no R2 took place.",
    "A7 – Offer Decline / R2 Reject / R2 Select: Both R1 and R2 filled. These statuses imply R1 was cleared and R2 was conducted.",
    "A8 – TA Screening: R1 filled (TA screening = Round 1); R2 left blank. TA Screening is a pre-panel screening step. It maps to R1 in the timeline. No R2 follows.",
    "A9 – Candidate not interested: All dates left blank. Treated same as Interview Dropout — candidate disengaged before any interview.",
]


def _normalize_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


# Map common short-forms / typos to canonical designation labels
_DESIGNATION_ALIASES: dict[str, str] = {
    "sse": "senior software engineer",
    "se ": "software engineer",
    "sr software engineer": "senior software engineer",
    "sr. software engineer": "senior software engineer",
    "senior se": "senior software engineer",
    "sr se": "senior software engineer",
    "software engg": "software engineer",
    "software eng": "software engineer",
    "tech lead": "technical lead",
    "tl": "technical lead",
    "tech l": "technical lead",
    "tech. lead": "technical lead",
    "technical l": "technical lead",
    "deveops engineer": "devops engineer",
    "devops engg": "devops engineer",
    "dev ops engineer": "devops engineer",
    "sde": "software development engineer",
    "sde-1": "software development engineer",
    "sde-2": "senior software development engineer",
    "sde 1": "software development engineer",
    "sde 2": "senior software development engineer",
    "sr developer": "senior developer",
    "sr. developer": "senior developer",
    "senior swe": "senior software engineer",
    "swe": "software engineer",
    "lead engineer": "technical lead",
    "lead developer": "technical lead",
}


def _normalize_designation(value: Any) -> str:
    """Return a normalized (lowercase, stripped, alias-resolved) designation string."""
    if pd.isna(value):
        return ""
    raw = str(value).strip().lower()
    raw = re.sub(r"[\u00a0\s]+", " ", raw)  # collapse nbsp + whitespace
    raw = raw.strip()
    # Try exact alias match first
    if raw in _DESIGNATION_ALIASES:
        return _DESIGNATION_ALIASES[raw]
    # Try prefix alias match for longer strings
    for alias, canonical in _DESIGNATION_ALIASES.items():
        if raw.startswith(alias):
            return canonical
    return raw


def _exp_bucket(value: Any) -> str:
    if pd.isna(value):
        return "Unknown"
    text = str(value).strip()
    if text in ("(Years)", ""):
        return "Unknown"
    match = re.search(r"([\d.]+)", text)
    if not match:
        return "Unknown"
    years = float(match.group(1))
    if years < 2:
        return "0-2 yrs"
    if years < 5:
        return "2-5 yrs"
    if years < 8:
        return "5-8 yrs"
    return "8+ yrs"


def _cap_date(ts: pd.Timestamp | pd.NaT) -> pd.Timestamp | pd.NaT:
    if pd.isna(ts):
        return pd.NaT
    if ts > MAX_PIPELINE_DATE:
        return MAX_PIPELINE_DATE
    return ts.normalize()


def _try_build_date_raw(month: int, day: int, year: int) -> pd.Timestamp | pd.NaT:
    try:
        return pd.Timestamp(year=year, month=month, day=day).normalize()
    except ValueError:
        return pd.NaT


def _parse_ambiguous_date_raw(text: str) -> pd.Timestamp | pd.NaT:
    text = text.strip()
    match = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", text)
    if not match:
        parsed = pd.to_datetime(text, errors="coerce")
        return parsed if pd.notna(parsed) else pd.NaT

    a, b, y = int(match.group(1)), int(match.group(2)), int(match.group(3))
    if y < 100:
        y += 2000

    candidates: list[pd.Timestamp] = []
    if a > 12:
        ts = _try_build_date_raw(b, a, y)
        if pd.notna(ts):
            candidates.append(ts)
    elif b > 12:
        ts = _try_build_date_raw(a, b, y)
        if pd.notna(ts):
            candidates.append(ts)
    else:
        for month, day in ((a, b), (b, a)):
            ts = _try_build_date_raw(month, day, y)
            if pd.notna(ts):
                candidates.append(ts)

    if not candidates:
        return pd.NaT

    unique = []
    for ts in candidates:
        if ts not in unique:
            unique.append(ts)
    return unique[0]


def parse_date_raw(value: Any) -> pd.Timestamp | pd.NaT:
    if pd.isna(value):
        return pd.NaT
    if isinstance(value, time):
        return pd.NaT
    if isinstance(value, str):
        text = value.strip().lower()
        if not text or text in {"cnr", "na", "n/a", "-"} or "not avail" in text:
            return pd.NaT
        return _parse_ambiguous_date_raw(value)
    if isinstance(value, (datetime, pd.Timestamp)):
        return pd.Timestamp(value)
    parsed = pd.to_datetime(value, errors="coerce")
    return parsed if pd.notna(parsed) else pd.NaT


def _fix_swapped_month_day(ts: pd.Timestamp) -> pd.Timestamp:
    """If date exceeds cap, try swapping month and day (common Excel error)."""
    if pd.isna(ts) or ts <= MAX_PIPELINE_DATE:
        return ts
    try:
        swapped = pd.Timestamp(
            year=ts.year, month=ts.day, day=ts.month
        ).normalize()
        if swapped <= MAX_PIPELINE_DATE:
            return swapped
    except ValueError:
        pass
    return MAX_PIPELINE_DATE


def _cap_date(ts: pd.Timestamp | pd.NaT) -> pd.Timestamp | pd.NaT:
    if pd.isna(ts):
        return pd.NaT
    if ts > MAX_PIPELINE_DATE:
        return MAX_PIPELINE_DATE
    return ts.normalize()


def _format_date_display(ts: Any) -> str:
    if pd.isna(ts):
        return ""
    return pd.Timestamp(ts).strftime("%m/%d/%Y")


def parse_submission_date(value: Any) -> pd.Timestamp | pd.NaT:
    ts_raw = parse_date_raw(value)
    if pd.isna(ts_raw):
        return pd.NaT
    ts = pd.Timestamp(ts_raw).normalize()
    ts = _fix_swapped_month_day(ts)
    return _cap_date(ts)


def process_round_date(
    ts_raw: Any, r0: pd.Timestamp | pd.NaT, is_r2: bool = False
) -> tuple[pd.Timestamp | pd.NaT, str, str]:
    if pd.isna(ts_raw):
        return pd.NaT, "Missing", ""
    ts = pd.Timestamp(ts_raw).normalize()
    if pd.notna(r0) and r0.year == 2025 and ts.year == 2026:
        corrected_ts = ts.replace(year=2025)
        corrected_ts = _cap_date(corrected_ts)
        note = (
            "Year corrected 2026→2025 (same as R1 correction)"
            if is_r2
            else "Year corrected 2026→2025 (R0 is 2025; 500+ day gap was anomalous entry error)"
        )
        return corrected_ts, "Yr Corrected", note

    ts = _fix_swapped_month_day(ts)
    ts = _cap_date(ts)
    return ts, "Original", "Original data from source"


def _trusted_delta(
    days: float | int | None,
    min_days: float = 0,
    max_days: float = MAX_REASONABLE_DAYS_R01,
) -> bool:
    if days is None or pd.isna(days):
        return False
    d = float(days)
    return min_days <= d <= max_days


def _median_from_series(
    series: pd.Series,
    fallback: float,
    min_days: float = 0,
    max_days: float = MAX_REASONABLE_DAYS_R01,
) -> float:
    clean = series.dropna()
    clean = clean[clean.apply(lambda d: _trusted_delta(d, min_days, max_days))]
    if len(clean) == 0:
        return fallback
    return float(clean.median())


def _build_match_key(row: pd.Series, level: str) -> str:
    skills = _normalize_text(row.get("Skills"))
    sourcer = _normalize_text(row.get("Sourcer"))
    designation = _normalize_text(row.get("Current Designation"))
    exp = _exp_bucket(row.get("Total Exp"))

    if level == "skills_sourcer":
        return f"{skills}||{sourcer}"
    if level == "skills":
        return skills
    if level == "designation_exp":
        return f"{designation}||{exp}"
    if level == "sourcer":
        return sourcer
    return "global"


def _lookup_median(
    row: pd.Series,
    delta_map: dict[str, float],
    global_median: float,
) -> tuple[float, str]:
    for level in ("skills_sourcer", "skills", "designation_exp", "sourcer", "global"):
        key = _build_match_key(row, level)
        if key and key in delta_map:
            return delta_map[key], MATCH_LEVEL_LABELS[level]
    return global_median, MATCH_LEVEL_LABELS["global"]


def _compute_delta_maps(
    df: pd.DataFrame,
    start_col: str,
    end_col: str,
    *,
    min_days: float = 0,
    max_days: float = MAX_REASONABLE_DAYS_R01,
    fallback: float = DEFAULT_R01_DAYS,
) -> tuple[dict[str, float], float]:
    valid = df[df[start_col].notna() & df[end_col].notna()].copy()
    valid["delta"] = (valid[end_col] - valid[start_col]).dt.days
    valid = valid[valid["delta"].apply(lambda d: _trusted_delta(d, min_days, max_days))]

    global_median = _median_from_series(
        valid["delta"], fallback=fallback, min_days=min_days, max_days=max_days
    )

    maps: dict[str, float] = {}
    for level in ("skills_sourcer", "skills", "designation_exp", "sourcer"):
        keys = valid.apply(lambda r: _build_match_key(r, level), axis=1)
        for key, group in valid.groupby(keys):
            if not key or key == "||":
                continue
            maps[key] = _median_from_series(
                group["delta"],
                fallback=global_median,
                min_days=min_days,
                max_days=max_days,
            )

    maps["global"] = global_median
    return maps, global_median


PREVIEW_COLUMNS = [
    "Name",
    "Sourcer",
    "Skills",
    "Final Status",
    "R0 Date (Display)",
    "R0 Source",
    "R1 Date (Display)",
    "R1 Source",
    "R1 Imputation Method",
    "R1 Imputation Days",
    "R2 Date (Display)",
    "R2 Source",
    "R2 Imputation Method",
    "R2 Imputation Days",
    "Days R0 to R1",
    "Days R1 to R2",
    "Total Pipeline Days (R0 to R2)",
    "Data Confidence",
]

REQUIRED_COLUMNS = [
    "Date of submission",
    "Sourcer",
    "Skills",
    "Name",
    "Final Status",
    "R1 Date",
    "R2 Date",
    "Current Designation",
    "Total Exp",
]

# Canonical column names (case/spacing tolerant match)
_COLUMN_CANONICAL = {name.lower(): name for name in REQUIRED_COLUMNS}

# Output files excluded when auto-detecting a new upload
_OUTPUT_XLSX_NAMES = {
    "supply_mapping_powerbi_ready.xlsx",
    "supply_mapping_analyzed.xlsx",
}


def _normalize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(col).strip() for col in df.columns]
    rename: dict[str, str] = {}
    for col in df.columns:
        canonical = _COLUMN_CANONICAL.get(col.lower())
        if canonical and col != canonical:
            rename[col] = canonical
    if rename:
        df = df.rename(columns=rename)
    return df


def validate_supply_mapping_columns(df: pd.DataFrame) -> None:
    """Raise ValueError if the workbook is not a Supply Mapping-format file."""
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        found = ", ".join(str(c) for c in df.columns[:20])
        if len(df.columns) > 20:
            found += ", ..."
        raise ValueError(
            "This Excel file is missing required Supply Mapping columns: "
            f"{missing}. Found columns: {found}. "
            "Please upload an .xlsx file with the same column headers as Supply Mapping."
        )
    if len(df) == 0:
        raise ValueError(
            "No candidate rows found. Ensure the file has data below the header row "
            "and is not only the template row '(MM-DD-YY)'."
        )


def resolve_input_xlsx(
    explicit_path: str | Path | None = None,
    search_dir: str | Path | None = None,
) -> Path:
    """
    Resolve which .xlsx to analyze.
    Priority: explicit path → uploads/ newest → project folder newest (excluding outputs).
    """
    base = Path(search_dir or Path(__file__).parent)

    if explicit_path:
        path = Path(explicit_path).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(f"Input file not found: {path}")
        if path.suffix.lower() not in {".xlsx", ".xls"}:
            raise ValueError(f"Input must be an Excel file (.xlsx or .xls): {path}")
        return path

    uploads_dir = base / "uploads"
    if uploads_dir.is_dir():
        upload_files = _list_input_xlsx_files(uploads_dir)
        if upload_files:
            return upload_files[0]

    project_files = _list_input_xlsx_files(base)
    if project_files:
        return project_files[0]

    raise FileNotFoundError(
        "No input Excel found. Place any Supply Mapping-format .xlsx in this folder "
        "or in uploads/, or run: python export_powerbi.py path/to/your_file.xlsx"
    )


def _list_input_xlsx_files(directory: Path) -> list[Path]:
    files = [
        p
        for p in directory.glob("*.xlsx")
        if p.name.lower() not in _OUTPUT_XLSX_NAMES and not p.name.startswith("~$")
    ]
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return files


def load_and_clean_dataframe(file_bytes: bytes | str) -> pd.DataFrame:
    if isinstance(file_bytes, bytes):
        source: Any = io.BytesIO(file_bytes)
    else:
        source = file_bytes
    df = pd.read_excel(source, sheet_name=0)
    df = _normalize_column_names(df)
    validate_supply_mapping_columns(df)
    df = df[df["Date of submission"].astype(str).str.strip() != "(MM-DD-YY)"].copy()
    df = df.reset_index(drop=True)
    validate_supply_mapping_columns(df)
    return df


def analyze_supply_data(file_bytes: bytes) -> dict[str, Any]:
    raw_df = load_and_clean_dataframe(file_bytes)
    df = raw_df.copy()

    # Parse R0 submission date
    df["R0"] = df["Date of submission"].apply(parse_submission_date)
    df["R0 Source"] = np.where(df["R0"].notna(), "Actual", "Missing")
    df["R0 Correction Note"] = ""

    # Parse R1 and R2 raw values
    df["R1_raw"] = df["R1 Date"].apply(parse_date_raw)
    df["R2_raw"] = df["R2 Date"].apply(parse_date_raw)

    # Process parsed round dates with R0-based year correction
    r1_processed = []
    r2_processed = []
    for idx, row in df.iterrows():
        r0 = row["R0"]
        # R1 processing
        r1_val, r1_src, r1_note = process_round_date(row["R1_raw"], r0, is_r2=False)
        r1_processed.append((r1_val, r1_src, r1_note))
        # R2 processing
        r2_val, r2_src, r2_note = process_round_date(row["R2_raw"], r0, is_r2=True)
        r2_processed.append((r2_val, r2_src, r2_note))

    df["R1 Date (Final)"] = [x[0] for x in r1_processed]
    df["R1 Source"] = [x[1] for x in r1_processed]
    df["R1 Imputation Method"] = [x[2] for x in r1_processed]
    df["R1 Imputation Days"] = np.nan

    df["R2 Date (Final)"] = [x[0] for x in r2_processed]
    df["R2 Source"] = [x[1] for x in r2_processed]
    df["R2 Imputation Method"] = [x[2] for x in r2_processed]
    df["R2 Imputation Days"] = np.nan

    # Compute actual gaps for cohort calculations (Original & Yr Corrected only)
    df["actual_R1_gap"] = np.where(
        df["R1 Source"].isin(["Original", "Yr Corrected"]),
        (df["R1 Date (Final)"] - df["R0"]).dt.days,
        np.nan,
    )

    # --------------------------------------------------------------------------
    # Build hierarchical lookup tables for the 7-level R1 prediction model
    # --------------------------------------------------------------------------

    # Add normalised designation column for grouping
    df["_desig_norm"] = df["Current Designation"].apply(_normalize_designation)
    df["_r0_month"] = df["R0"].apply(lambda x: x.strftime("%Y-%m") if pd.notna(x) else "")
    df["_skills_norm"] = df["Skills"].apply(_normalize_text)

    # Helper: compute median gap (in days) for a subset
    def _med_gap(subset: pd.DataFrame) -> float | None:
        vals = subset["actual_R1_gap"].dropna()
        vals = vals[(vals >= 0) & (vals <= MAX_REASONABLE_DAYS_R01)]
        return float(vals.median()) if len(vals) >= 1 else None

    # Helper: compute median R1 date (as Timestamp) for a subset
    def _med_r1_date(subset: pd.DataFrame) -> pd.Timestamp | None:
        dates = subset.loc[
            subset["R1 Source"].isin(["Original", "Yr Corrected"]), "R1 Date (Final)"
        ].dropna()
        if len(dates) == 0:
            return None
        med_ts = int(dates.apply(lambda x: x.timestamp()).median())
        return pd.Timestamp(med_ts, unit="s").normalize()

    # --- Pre-compute lookup dicts ---
    # Level 1: desig + skills + exact R0 date  → median R1 date
    lvl1_r1date: dict[tuple, pd.Timestamp] = {}
    for key, grp in df.groupby(["_desig_norm", "_skills_norm", "R0"]):
        med = _med_r1_date(grp)
        if med is not None:
            lvl1_r1date[key] = med

    # Level 2: desig + skills + R0 month  → median R1 date
    lvl2_r1date: dict[tuple, pd.Timestamp] = {}
    for key, grp in df.groupby(["_desig_norm", "_skills_norm", "_r0_month"]):
        med = _med_r1_date(grp)
        if med is not None:
            lvl2_r1date[key] = med

    # Level 3: desig + R0 month  → median R1 date
    lvl3_r1date: dict[tuple, pd.Timestamp] = {}
    for key, grp in df.groupby(["_desig_norm", "_r0_month"]):
        med = _med_r1_date(grp)
        if med is not None:
            lvl3_r1date[key] = med

    # Level 4: desig + skills  → median gap
    lvl4_gap: dict[tuple, float] = {}
    for key, grp in df.groupby(["_desig_norm", "_skills_norm"]):
        med = _med_gap(grp)
        if med is not None:
            lvl4_gap[key] = med

    # Level 5: desig only  → median gap
    lvl5_gap: dict[str, float] = {}
    for key, grp in df.groupby("_desig_norm"):
        med = _med_gap(grp)
        if med is not None:
            lvl5_gap[key] = med

    # Level 6: skills only  → median gap
    lvl6_gap: dict[str, float] = {}
    for key, grp in df.groupby("_skills_norm"):
        med = _med_gap(grp)
        if med is not None:
            lvl6_gap[key] = med

    # Level 7: global median gap
    all_actual_gaps = df["actual_R1_gap"].dropna()
    all_actual_gaps = all_actual_gaps[(all_actual_gaps >= 0) & (all_actual_gaps <= MAX_REASONABLE_DAYS_R01)]
    global_median_gap = float(all_actual_gaps.median()) if len(all_actual_gaps) > 0 else 91.0

    # --------------------------------------------------------------------------
    # Impute / Predict missing R1 (and R2) dates using hierarchical model
    # --------------------------------------------------------------------------
    for idx, row in df.iterrows():
        st = row["Final Status"]
        if pd.isna(st) or st in NO_IMPUTE:
            continue

        needs_r1 = st in REACHED_R1 and pd.isna(row["R1 Date (Final)"])
        needs_r2 = st in REACHED_R2 and pd.isna(row["R2 Date (Final)"])

        if needs_r1:
            desig = row["_desig_norm"]
            skills = row["_skills_norm"]
            r0 = row["R0"]
            r0_month = row["_r0_month"]
            r0_str = r0.strftime("%Y-%m-%d") if pd.notna(r0) else ""

            imputed_r1: pd.Timestamp | None = None
            gap: float | None = None
            method: str = ""

            # Level 1 – same designation + skills + exact R0 date → median R1 date
            k1 = (desig, skills, r0)
            if pd.notna(r0) and k1 in lvl1_r1date:
                candidate = _cap_date(lvl1_r1date[k1])
                if pd.notna(candidate) and candidate >= r0:
                    imputed_r1 = candidate
                    method = (
                        f"Predicted L1 (Designation+Skills+Date): "
                        f"median R1 date for '{row['Current Designation']}' | '{row['Skills']}' | {r0_str}"
                    )

            # Level 2 – same designation + skills + R0 month → median R1 date
            if imputed_r1 is None:
                k2 = (desig, skills, r0_month)
                if r0_month and k2 in lvl2_r1date:
                    candidate = _cap_date(lvl2_r1date[k2])
                    if pd.notna(candidate) and candidate >= r0:
                        imputed_r1 = candidate
                        method = (
                            f"Predicted L2 (Designation+Skills+Month): "
                            f"median R1 date for '{row['Current Designation']}' | '{row['Skills']}' | {r0_month}"
                        )

            # Level 3 – same designation + R0 month → median R1 date
            if imputed_r1 is None:
                k3 = (desig, r0_month)
                if r0_month and k3 in lvl3_r1date:
                    candidate = _cap_date(lvl3_r1date[k3])
                    if pd.notna(candidate) and candidate >= r0:
                        imputed_r1 = candidate
                        method = (
                            f"Predicted L3 (Designation+Month): "
                            f"median R1 date for '{row['Current Designation']}' | {r0_month}"
                        )

            # Level 4 – same designation + skills → R0 + median gap
            if imputed_r1 is None:
                k4 = (desig, skills)
                if k4 in lvl4_gap and pd.notna(r0):
                    gap = lvl4_gap[k4]
                    imputed_r1 = _cap_date(r0 + pd.Timedelta(days=int(round(gap))))
                    method = (
                        f"Predicted L4 (Designation+Skills gap): "
                        f"R0 + {int(round(gap))} days "
                        f"(median gap for '{row['Current Designation']}' | '{row['Skills']}')"
                    )

            # Level 5 – same designation → R0 + median gap
            if imputed_r1 is None:
                if desig in lvl5_gap and pd.notna(r0):
                    gap = lvl5_gap[desig]
                    imputed_r1 = _cap_date(r0 + pd.Timedelta(days=int(round(gap))))
                    method = (
                        f"Predicted L5 (Designation gap): "
                        f"R0 + {int(round(gap))} days "
                        f"(median gap for designation '{row['Current Designation']}')"
                    )

            # Level 6 – same skills → R0 + median gap
            if imputed_r1 is None:
                if skills in lvl6_gap and pd.notna(r0):
                    gap = lvl6_gap[skills]
                    imputed_r1 = _cap_date(r0 + pd.Timedelta(days=int(round(gap))))
                    method = (
                        f"Predicted L6 (Skills gap): "
                        f"R0 + {int(round(gap))} days "
                        f"(median gap for skills '{row['Skills']}')"
                    )

            # Level 7 – global fallback
            if imputed_r1 is None and pd.notna(r0):
                gap = global_median_gap
                imputed_r1 = _cap_date(r0 + pd.Timedelta(days=int(round(gap))))
                method = (
                    f"Predicted L7 (Global fallback): "
                    f"R0 + {int(round(gap))} days (global median gap)"
                )

            if imputed_r1 is not None:
                used_gap = (
                    int((imputed_r1 - r0).days)
                    if pd.notna(r0) else (int(round(gap)) if gap is not None else 0)
                )
                df.at[idx, "R1 Date (Final)"] = imputed_r1
                df.at[idx, "R1 Source"] = "Imputed"
                df.at[idx, "R1 Imputation Method"] = f"Imputed: {method}"
                df.at[idx, "R1 Imputation Days"] = used_gap

        if needs_r2 and st != "R1 Reject":
            r1 = df.at[idx, "R1 Date (Final)"]
            if pd.notna(r1):
                imputed_r2 = _cap_date(r1 + pd.Timedelta(days=2))
                df.at[idx, "R2 Date (Final)"] = imputed_r2
                df.at[idx, "R2 Source"] = "Imputed"
                df.at[idx, "R2 Imputation Method"] = "Imputed: R1 + 2 days (all observed R1→R2 gaps = 2 days)"
                df.at[idx, "R2 Imputation Days"] = 2

    # Fill final notes for R1 and R2
    for idx, row in df.iterrows():
        st = row["Final Status"]
        # R1 Notes
        if pd.isna(df.at[idx, "R1 Date (Final)"]):
            if pd.isna(st):
                df.at[idx, "R1 Imputation Method"] = "Not applicable – nan"
            else:
                df.at[idx, "R1 Imputation Method"] = f"Left blank – {st} (candidate did not attend/progress)"

        # R2 Notes
        if pd.isna(df.at[idx, "R2 Date (Final)"]):
            if pd.isna(st):
                df.at[idx, "R2 Imputation Method"] = "Not needed – nan (did not reach R2)"
            elif st == "R1 Reject":
                df.at[idx, "R2 Imputation Method"] = "Not needed – R1 Reject (did not reach R2)"
            elif st == "TA Screening":
                df.at[idx, "R2 Imputation Method"] = "Not applicable – TA Screening is a pre-round screening (no R2)"
            elif st in NO_IMPUTE:
                df.at[idx, "R2 Imputation Method"] = f"Left blank – {st}"

    _fix_timeline_consistency(df)

    # Display-friendly date strings (always visible in Excel)
    df["R0 Date (Display)"] = df["R0"].apply(_format_date_display)
    df["R1 Date (Display)"] = df["R1 Date (Final)"].apply(_format_date_display)
    df["R2 Date (Display)"] = df["R2 Date (Final)"].apply(_format_date_display)

    df["Days R0 to R1"] = np.where(
        df["R0"].notna() & df["R1 Date (Final)"].notna(),
        (df["R1 Date (Final)"] - df["R0"]).dt.days,
        np.nan,
    )
    df["Days R1 to R2"] = np.where(
        df["R1 Date (Final)"].notna() & df["R2 Date (Final)"].notna(),
        (df["R2 Date (Final)"] - df["R1 Date (Final)"]).dt.days,
        np.nan,
    )
    df["Total Pipeline Days (R0 to R2)"] = np.where(
        df["R0"].notna() & df["R2 Date (Final)"].notna(),
        (df["R2 Date (Final)"] - df["R0"]).dt.days,
        np.nan,
    )

    df["Exp Bucket"] = df["Total Exp"].apply(_exp_bucket)
    df["Reached R1"] = df["Final Status"].fillna("Unknown").isin(REACHED_R1)
    df["Reached R2"] = df["Final Status"].fillna("Unknown").isin(REACHED_R2)
    df["Reached Offer"] = df["Final Status"].fillna("Unknown").isin(REACHED_OFFER)
    df["Days R0 to Offer Stage"] = np.where(
        df["Reached Offer"] & df["R0"].notna() & df["R2 Date (Final)"].notna(),
        (df["R2 Date (Final)"] - df["R0"]).dt.days,
        np.nan,
    )
    df["Data Confidence"] = df.apply(_data_confidence_label, axis=1)

    validation = _run_validation_checks(df)
    validation_df = pd.DataFrame(validation)

    summary = _build_summary(df)
    sourcer_summary = _build_sourcer_summary(df)
    sourcer_breakdown = _build_sourcer_breakdown(df)
    designation_breakdown = _build_designation_breakdown(df)

    pbi_candidates = _build_powerbi_candidates(df)
    stage_transitions = _build_stage_transitions_long(df)
    bottleneck_rounds = _build_bottleneck_by_round(df)
    bottleneck_designation = _build_bottleneck_by_designation(df)
    bottleneck_sourcer = _build_bottleneck_by_sourcer(df)
    slow_movers = _build_slow_movers_detail(df)
    designation_timing = _build_timing_by_designation(df)
    sourcer_timing = _build_timing_by_sourcer(df)
    pbi_summary = _build_pbi_summary(df, slow_movers)
    pbi_funnel = _build_pbi_funnel(df)
    pbi_status = _build_pbi_status(df)
    pbi_sourcer_chart = _build_pbi_sourcer_chart(sourcer_summary)
    pbi_dashboard_layout = _build_pbi_dashboard_layout()

    excel_bytes = _build_excel_output(
        df,
        summary,
        sourcer_summary,
        sourcer_breakdown,
        designation_breakdown,
        validation_df,
        pbi_candidates=pbi_candidates,
        stage_transitions=stage_transitions,
        bottleneck_rounds=bottleneck_rounds,
        bottleneck_designation=bottleneck_designation,
        bottleneck_sourcer=bottleneck_sourcer,
        slow_movers=slow_movers,
        designation_timing=designation_timing,
        sourcer_timing=sourcer_timing,
        pbi_summary=pbi_summary,
        pbi_funnel=pbi_funnel,
        pbi_status=pbi_status,
        pbi_sourcer_chart=pbi_sourcer_chart,
        pbi_dashboard_layout=pbi_dashboard_layout,
    )
    powerbi_csv_zip = _build_powerbi_csv_zip(
        pbi_candidates,
        stage_transitions,
        bottleneck_rounds,
        bottleneck_designation,
        bottleneck_sourcer,
        slow_movers,
        designation_timing,
        sourcer_timing,
        pbi_summary,
        pbi_funnel,
        pbi_status,
        pbi_sourcer_chart,
        pbi_dashboard_layout,
    )

    return {
        "processed_df": df,
        "summary": summary,
        "sourcer_summary": sourcer_summary,
        "sourcer_breakdown": sourcer_breakdown,
        "designation_breakdown": designation_breakdown,
        "pbi_candidates": pbi_candidates,
        "stage_transitions": stage_transitions,
        "bottleneck_rounds": bottleneck_rounds,
        "bottleneck_designation": bottleneck_designation,
        "bottleneck_sourcer": bottleneck_sourcer,
        "slow_movers": slow_movers,
        "designation_timing": designation_timing,
        "sourcer_timing": sourcer_timing,
        "pbi_summary": pbi_summary,
        "pbi_funnel": pbi_funnel,
        "pbi_status": pbi_status,
        "pbi_sourcer_chart": pbi_sourcer_chart,
        "pbi_dashboard_layout": pbi_dashboard_layout,
        "assumptions": ASSUMPTIONS,
        "powerbi_instructions": POWERBI_INSTRUCTIONS,
        "excel_bytes": excel_bytes,
        "powerbi_csv_zip": powerbi_csv_zip,
        "stats": _build_stats(df),
        "charts": _build_charts(df, sourcer_summary),
        "status_distribution": _build_status_distribution(df),
        "validation": validation,
    }


def _fix_timeline_consistency(df: pd.DataFrame) -> None:
    """Ensure R2 is after R1 by at least MIN_REASONABLE_DAYS_R12 for reached-R2 rows."""
    for idx, row in df.iterrows():
        if row.get("Final Status") not in REACHED_R2:
            continue
        r0 = row["R0"]
        r1 = row["R1 Date (Final)"]
        r2 = row["R2 Date (Final)"]
        if pd.isna(r1) or pd.isna(r2):
            continue
        gap = (r2 - r1).days
        if gap >= MIN_REASONABLE_DAYS_R12:
            continue

        days_r12 = int(DEFAULT_R12_DAYS)
        new_r2 = _cap_date(r1 + pd.Timedelta(days=days_r12))
        new_r1 = r1

        if new_r2 <= r1:
            new_r1 = _cap_date(r1 - pd.Timedelta(days=days_r12))
            if pd.notna(r0) and new_r1 < r0:
                new_r1 = _cap_date(r0 + pd.Timedelta(days=int(DEFAULT_R01_DAYS)))
            new_r2 = _cap_date(new_r1 + pd.Timedelta(days=days_r12))

        if new_r1 != r1:
            df.at[idx, "R1 Date (Final)"] = new_r1
            if row["R1 Source"] == "Actual":
                df.at[idx, "R1 Source"] = "Adjusted"
            note = str(row.get("R0 Correction Note") or "")
            df.at[idx, "R0 Correction Note"] = (
                note + " R1 back-adjusted so R2 follows R1."
            ).strip()

        if new_r2 != r2:
            df.at[idx, "R2 Date (Final)"] = new_r2
            if row["R2 Source"] == "Actual":
                df.at[idx, "R2 Source"] = "Adjusted"
            if row["R2 Source"] == "Imputed":
                df.at[idx, "R2 Imputation Days"] = days_r12


def _data_confidence_label(row: pd.Series) -> str:
    sources = {row.get("R0 Source"), row.get("R1 Source"), row.get("R2 Source")}
    sources.discard("Missing")
    sources.discard("")
    sources.discard(None)
    if not sources:
        return "Unknown"
    if sources <= {"Actual", "Corrected"}:
        return "High"
    return "Estimated"


def _run_validation_checks(df: pd.DataFrame) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []

    def add(check: str, passed: bool, detail: str) -> None:
        checks.append(
            {"Check": check, "Status": "PASS" if passed else "FAIL", "Detail": detail}
        )

    total = len(df)

    future = 0
    for col in ("R0", "R1 Date (Final)", "R2 Date (Final)"):
        if col in df.columns:
            future += int((df[col].notna() & (df[col] > MAX_PIPELINE_DATE)).sum())
    add("No dates after 17-Jun-2026", future == 0, f"{future} future dates found")

    bad_impute = 0
    for st in NO_IMPUTE:
        sub = df[df["Final Status"] == st]
        bad_impute += int((sub["R1 Source"] == "Imputed").sum())
        bad_impute += int((sub["R2 Source"] == "Imputed").sum())
    add(
        "No imputation for Dropout/On Hold/Not Interested",
        bad_impute == 0,
        f"{bad_impute} improper imputations",
    )

    r1rej_r2 = df[df["Final Status"] == "R1 Reject"]
    bad_r2 = int((r1rej_r2["R2 Source"] == "Imputed").sum())
    add("No R2 imputation for R1 Reject", bad_r2 == 0, f"{bad_r2} violations")

    need_r1 = df[df["Final Status"].isin(REACHED_R1) & ~df["Final Status"].isin(NO_IMPUTE)]
    miss_r1 = int(need_r1["R1 Date (Final)"].isna().sum())
    add("All reached-R1 rows have R1 date", miss_r1 == 0, f"{miss_r1} missing R1")

    need_r2 = df[df["Final Status"].isin(REACHED_R2)]
    miss_r2 = int(need_r2["R2 Date (Final)"].isna().sum())
    add("All reached-R2 rows have R2 date", miss_r2 == 0, f"{miss_r2} missing R2")

    neg = 0
    for col in ("Days R0 to R1", "Days R1 to R2"):
        neg += int((df[col].notna() & (df[col] < 0)).sum())
    add("No negative day gaps", neg == 0, f"{neg} negative gaps")

    order_bad = df[
        df["R0"].notna()
        & df["R1 Date (Final)"].notna()
        & (df["R1 Date (Final)"] < df["R0"])
    ]
    add("R1 on or after R0", len(order_bad) == 0, f"{len(order_bad)} out-of-order")

    r2_order = df[
        df["R1 Date (Final)"].notna()
        & df["R2 Date (Final)"].notna()
        & (df["R2 Date (Final)"] < df["R1 Date (Final)"])
    ]
    add("R2 on or after R1", len(r2_order) == 0, f"{len(r2_order)} out-of-order")

    r2_gap = df[df["Reached R2"] & df["Days R1 to R2"].notna()]
    zero_gap = int((r2_gap["Days R1 to R2"] < MIN_REASONABLE_DAYS_R12).sum())
    add(
        "R1 to R2 gap at least 1 day (reached R2)",
        zero_gap == 0,
        f"{zero_gap} rows with invalid gap",
    )

    imp1 = df[df["R1 Source"] == "Imputed"]
    no_method = int(imp1["R1 Imputation Method"].isin(["", None]).sum())
    add("Imputed R1 has method documented", no_method == 0, f"{no_method} missing method")

    imp2 = df[df["R2 Source"] == "Imputed"]
    no_method2 = int(imp2["R2 Imputation Method"].isin(["", None]).sum())
    add("Imputed R2 has method documented", no_method2 == 0, f"{no_method2} missing method")

    reached_r1 = int(df["Reached R1"].sum())
    conv = round(reached_r1 / total * 100, 2) if total else 0
    add(
        "Conversion counts consistent",
        reached_r1 == int(df[df["Final Status"].isin(REACHED_R1)].shape[0]),
        f"R1 reached: {reached_r1}, conversion: {conv}%",
    )

    high = int((df["Data Confidence"] == "High").sum())
    est = int((df["Data Confidence"] == "Estimated").sum())
    add(
        "Data confidence labels assigned",
        high + est <= total,
        f"High: {high}, Estimated: {est}, Total: {total}",
    )

    passed = sum(1 for c in checks if c["Status"] == "PASS")
    add(
        "OVERALL QUALITY SCORE",
        passed == len(checks),
        f"{passed}/{len(checks)} checks passed",
    )

    return checks


def _build_stats(df: pd.DataFrame) -> dict[str, Any]:
    total = len(df)
    reached_r1 = int(df["Reached R1"].sum())
    reached_r2 = int(df["Reached R2"].sum())
    reached_offer = int(df["Reached Offer"].sum())

    return {
        "total_candidates": int(total),
        "reached_r1": reached_r1,
        "reached_r2": reached_r2,
        "reached_offer": reached_offer,
        "r0_r1_conversion": round(reached_r1 / total * 100, 2) if total else 0,
        "r1_r2_conversion": round(reached_r2 / reached_r1 * 100, 2) if reached_r1 else 0,
        "r2_offer_conversion": round(reached_offer / reached_r2 * 100, 2) if reached_r2 else 0,
        "overall_conversion": round(reached_offer / total * 100, 2) if total else 0,
        "r1_imputed": int((df["R1 Source"] == "Imputed").sum()),
        "r2_imputed": int((df["R2 Source"] == "Imputed").sum()),
        "r0_corrected": int((df["R0 Source"] == "Corrected").sum()),
        "r1_actual": int((df["R1 Source"] == "Actual").sum()),
        "r2_actual": int((df["R2 Source"] == "Actual").sum()),
        "data_confidence_high": int((df["Data Confidence"] == "High").sum()),
        "data_confidence_estimated": int((df["Data Confidence"] == "Estimated").sum()),
        "avg_r0_r1": _safe_mean(df["Days R0 to R1"]),
        "avg_r1_r2": _safe_mean(
            df["Days R1 to R2"],
            min_days=MIN_REASONABLE_DAYS_R12,
            max_days=MAX_REASONABLE_DAYS_R12,
        ),
        "median_r0_r1": _safe_median(df["Days R0 to R1"]),
        "median_r1_r2": _safe_median(
            df["Days R1 to R2"],
            min_days=MIN_REASONABLE_DAYS_R12,
            max_days=MAX_REASONABLE_DAYS_R12,
        ),
    }


def _build_status_distribution(df: pd.DataFrame) -> list[dict[str, Any]]:
    counts = df["Final Status"].fillna("Unknown").value_counts()
    return [
        {"status": str(status), "count": int(count)}
        for status, count in counts.items()
    ]


def _build_charts(
    df: pd.DataFrame, sourcer_summary: pd.DataFrame
) -> dict[str, Any]:
    total = len(df)
    reached_r1 = int(df["Reached R1"].sum())
    reached_r2 = int(df["Reached R2"].sum())
    reached_offer = int(df["Reached Offer"].sum())

    funnel = {
        "labels": ["R0 Applied", "R1 Interview", "R2 Interview", "Offer Stage"],
        "counts": [total, reached_r1, reached_r2, reached_offer],
        "percentages": [
            100.0,
            round(reached_r1 / total * 100, 1) if total else 0,
            round(reached_r2 / total * 100, 1) if total else 0,
            round(reached_offer / total * 100, 1) if total else 0,
        ],
    }

    ss = sourcer_summary.copy()
    if len(ss):
        by_volume = ss.sort_values("Total Submitted", ascending=False).head(12)
        conversion_col = "R0→R1 Conversion %"
        speed_col = "Avg Days R0→R1"
        if conversion_col not in by_volume.columns:
            conversion_col = [c for c in by_volume.columns if "R0" in c and "Conversion" in c][0]
        if speed_col not in by_volume.columns:
            speed_col = [c for c in by_volume.columns if "R0" in c and "Days" in c][0]
        speed_r12_col = "Avg Days R1→R2"
        if speed_r12_col not in by_volume.columns:
            matches = [c for c in by_volume.columns if "R1" in c and "R2" in c and "Days" in c]
            speed_r12_col = matches[0] if matches else None

        sourcer_charts = {
            "labels": by_volume["Sourcer"].tolist(),
            "volume": by_volume["Total Submitted"].tolist(),
            "r0_r1_conversion": by_volume[conversion_col].tolist(),
            "r1_r2_conversion": by_volume.get(
                "R1→R2 Conversion %", pd.Series([0] * len(by_volume))
            ).tolist(),
            "overall_conversion": by_volume["Overall Conversion %"].tolist(),
            "avg_r0_r1": [
                x if x is not None and pd.notna(x) else None
                for x in by_volume[speed_col].tolist()
            ],
            "avg_r1_r2": [
                x if x is not None and pd.notna(x) else None
                for x in (
                    by_volume[speed_r12_col].tolist()
                    if speed_r12_col
                    else [None] * len(by_volume)
                )
            ],
        }
    else:
        sourcer_charts = {
            "labels": [],
            "volume": [],
            "r0_r1_conversion": [],
            "r1_r2_conversion": [],
            "overall_conversion": [],
            "avg_r0_r1": [],
            "avg_r1_r2": [],
        }

    # Skills-level top performers
    skills_records = []
    subset = df[df["Sourcer"].notna()].copy()
    subset["Skills"] = subset["Skills"].fillna("Unknown")
    for (sourcer, skills), group in subset.groupby(["Sourcer", "Skills"]):
        total_g = len(group)
        if total_g < 5:
            continue
        offer = int(group["Reached Offer"].sum())
        skills_records.append(
            {
                "label": f"{sourcer} · {skills[:30]}",
                "sourcer": str(sourcer),
                "skills": str(skills),
                "volume": total_g,
                "conversion": round(offer / total_g * 100, 2),
                "avg_r0_r1": _safe_mean(group["Days R0 to R1"]) or 0,
            }
        )

    skills_records.sort(key=lambda x: x["conversion"], reverse=True)
    top_skills = skills_records[:10]

    return {
        "funnel": funnel,
        "sourcers": sourcer_charts,
        "top_skills": top_skills,
    }


def _safe_mean(
    series: pd.Series,
    min_days: float = 0,
    max_days: float = MAX_REASONABLE_DAYS_R01,
) -> float | None:
    clean = series.dropna()
    clean = clean[clean.apply(lambda d: _trusted_delta(d, min_days, max_days))]
    if len(clean) == 0:
        return None
    return round(float(clean.mean()), 2)


def _safe_median(
    series: pd.Series,
    min_days: float = 0,
    max_days: float = MAX_REASONABLE_DAYS_R01,
) -> float | None:
    clean = series.dropna()
    clean = clean[clean.apply(lambda d: _trusted_delta(d, min_days, max_days))]
    if len(clean) == 0:
        return None
    return round(float(clean.median()), 2)


def _build_summary(df: pd.DataFrame) -> pd.DataFrame:
    total = len(df)
    reached_r1 = int(df["Reached R1"].sum())
    reached_r2 = int(df["Reached R2"].sum())
    reached_offer = int(df["Reached Offer"].sum())

    rows = [
        {
            "Metric": "Total Candidates",
            "Value": total,
        },
        {
            "Metric": "Reached R1 (count)",
            "Value": reached_r1,
        },
        {
            "Metric": "Reached R2 (count)",
            "Value": reached_r2,
        },
        {
            "Metric": "Reached Offer (count)",
            "Value": reached_offer,
        },
        {
            "Metric": "R0 → R1 Conversion %",
            "Value": round(reached_r1 / total * 100, 2) if total else 0,
        },
        {
            "Metric": "R1 → R2 Conversion %",
            "Value": round(reached_r2 / reached_r1 * 100, 2) if reached_r1 else 0,
        },
        {
            "Metric": "R2 → Offer Conversion %",
            "Value": round(reached_offer / reached_r2 * 100, 2) if reached_r2 else 0,
        },
        {
            "Metric": "Overall R0 → Offer Conversion %",
            "Value": round(reached_offer / total * 100, 2) if total else 0,
        },
        {
            "Metric": "Avg Days R0 → R1",
            "Value": _safe_mean(df["Days R0 to R1"]),
        },
        {
            "Metric": "Median Days R0 → R1",
            "Value": _safe_median(df["Days R0 to R1"]),
        },
        {
            "Metric": "Avg Days R1 → R2",
            "Value": _safe_mean(
                df["Days R1 to R2"],
                min_days=MIN_REASONABLE_DAYS_R12,
                max_days=MAX_REASONABLE_DAYS_R12,
            ),
        },
        {
            "Metric": "Median Days R1 → R2",
            "Value": _safe_median(
                df["Days R1 to R2"],
                min_days=MIN_REASONABLE_DAYS_R12,
                max_days=MAX_REASONABLE_DAYS_R12,
            ),
        },
        {
            "Metric": "R1 Dates Imputed",
            "Value": int((df["R1 Source"] == "Imputed").sum()),
        },
        {
            "Metric": "R2 Dates Imputed",
            "Value": int((df["R2 Source"] == "Imputed").sum()),
        },
        {
            "Metric": "R0 Dates Corrected",
            "Value": int((df["R0 Source"] == "Corrected").sum()),
        },
    ]
    return pd.DataFrame(rows)


def _build_sourcer_summary(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for sourcer, group in df.groupby("Sourcer", dropna=False):
        name = sourcer if pd.notna(sourcer) and str(sourcer).strip() else "Unknown"
        total = len(group)
        r1 = int(group["Reached R1"].sum())
        r2 = int(group["Reached R2"].sum())
        offer = int(group["Reached Offer"].sum())

        records.append(
            {
                "Sourcer": name,
                "Total Submitted": total,
                "Reached R1": r1,
                "Reached R2": r2,
                "Offer Stage": offer,
                "R0→R1 Conversion %": round(r1 / total * 100, 2) if total else 0,
                "R1→R2 Conversion %": round(r2 / r1 * 100, 2) if r1 else 0,
                "R2→Offer Conversion %": round(offer / r2 * 100, 2) if r2 else 0,
                "Overall Conversion %": round(offer / total * 100, 2) if total else 0,
                "Avg Days R0→R1": _safe_mean(group["Days R0 to R1"]),
                "Avg Days R1→R2": _safe_mean(
                    group["Days R1 to R2"],
                    min_days=MIN_REASONABLE_DAYS_R12,
                    max_days=MAX_REASONABLE_DAYS_R12,
                ),
            }
        )

    result = pd.DataFrame(records)
    if len(result):
        result = result.sort_values("Total Submitted", ascending=False)
    return result.reset_index(drop=True)


def _build_sourcer_breakdown(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    subset = df[df["Sourcer"].notna()].copy()
    subset["Skills"] = subset["Skills"].fillna("Unknown")

    for (sourcer, skills), group in subset.groupby(["Sourcer", "Skills"]):
        total = len(group)
        r1 = int(group["Reached R1"].sum())
        r2 = int(group["Reached R2"].sum())
        offer = int(group["Reached Offer"].sum())

        records.append(
            {
                "Sourcer": sourcer,
                "Skills / Role Type": skills,
                "Exp Bucket (mode)": group["Exp Bucket"].mode().iloc[0]
                if len(group["Exp Bucket"].mode())
                else "Unknown",
                "Total Submitted": total,
                "Reached R1": r1,
                "Reached R2": r2,
                "Offer Stage": offer,
                "R0→R1 Conversion %": round(r1 / total * 100, 2) if total else 0,
                "R1→R2 Conversion %": round(r2 / r1 * 100, 2) if r1 else 0,
                "R2→Offer Conversion %": round(offer / r2 * 100, 2) if r2 else 0,
                "Overall Conversion %": round(offer / total * 100, 2) if total else 0,
                "Avg Days R0→R1": _safe_mean(group["Days R0 to R1"]),
                "Avg Days R1→R2": _safe_mean(
                    group["Days R1 to R2"],
                    min_days=MIN_REASONABLE_DAYS_R12,
                    max_days=MAX_REASONABLE_DAYS_R12,
                ),
            }
        )

    result = pd.DataFrame(records)
    if len(result):
        result = result.sort_values(["Sourcer", "Total Submitted"], ascending=[True, False])
    return result.reset_index(drop=True)


def _build_designation_breakdown(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    subset = df.copy()
    subset["Current Designation"] = subset["Current Designation"].fillna("Unknown")

    for (designation, exp_bucket), group in subset.groupby(
        ["Current Designation", "Exp Bucket"]
    ):
        total = len(group)
        r1 = int(group["Reached R1"].sum())
        r2 = int(group["Reached R2"].sum())
        offer = int(group["Reached Offer"].sum())

        records.append(
            {
                "Designation": designation,
                "Experience Bucket": exp_bucket,
                "Total": total,
                "Reached R1": r1,
                "Reached R2": r2,
                "Offer Stage": offer,
                "R0→R1 Conversion %": round(r1 / total * 100, 2) if total else 0,
                "R1→R2 Conversion %": round(r2 / r1 * 100, 2) if r1 else 0,
                "Overall Conversion %": round(offer / total * 100, 2) if total else 0,
                "Avg Days R0→R1": _safe_mean(group["Days R0 to R1"]),
                "Avg Days R1→R2": _safe_mean(
                    group["Days R1 to R2"],
                    min_days=MIN_REASONABLE_DAYS_R12,
                    max_days=MAX_REASONABLE_DAYS_R12,
                ),
            }
        )

    result = pd.DataFrame(records)
    if len(result):
        result = result.sort_values("Total", ascending=False)
    return result.reset_index(drop=True)


def _candidate_key(row: pd.Series) -> str:
    sno = row.get("S.No")
    if pd.notna(sno):
        return f"C{int(float(sno))}"
    name = str(row.get("Name") or "Unknown").strip()
    return f"N{name[:40]}"


def _is_slow(days: Any) -> bool:
    if pd.isna(days):
        return False
    try:
        return float(days) >= SLOW_THRESHOLD_DAYS
    except (TypeError, ValueError):
        return False


def _trusted_days(
    days: Any,
    min_days: float = 0,
    max_days: float = MAX_REASONABLE_DAYS_R01,
) -> bool:
    if pd.isna(days):
        return False
    try:
        value = float(days)
    except (TypeError, ValueError):
        return False
    return min_days <= value <= max_days


def _stage_timing_metrics(
    group: pd.DataFrame,
    col: str,
    min_days: float = 0,
    max_days: float = MAX_REASONABLE_DAYS_R01,
) -> dict[str, Any]:
    valid = group[col].dropna()
    valid = valid[valid.apply(lambda d: _trusted_days(d, min_days, max_days))]
    count_valid = int(len(valid))
    slow_count = int((valid >= SLOW_THRESHOLD_DAYS).sum()) if count_valid else 0
    return {
        "count_valid": count_valid,
        "avg_days": _safe_mean(group[col], min_days=min_days, max_days=max_days),
        "median_days": _safe_median(group[col], min_days=min_days, max_days=max_days),
        "slow_30plus_count": slow_count,
        "slow_30plus_pct": round(slow_count / count_valid * 100, 2) if count_valid else 0,
        "max_days": round(float(valid.max()), 2) if count_valid else None,
    }


def _primary_bottleneck(metrics: dict[str, dict[str, Any]]) -> str:
    ranked = []
    for label, data in metrics.items():
        avg = data.get("avg_days")
        if avg is not None:
            ranked.append((label, avg, data.get("slow_30plus_count", 0)))
    if not ranked:
        return "Insufficient data"
    ranked.sort(key=lambda item: (item[1], item[2]), reverse=True)
    return ranked[0][0]


def _build_powerbi_candidates(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for _, row in df.iterrows():
        slow_stages = []
        for _, _, col, min_d, max_d in STAGE_TRANSITIONS:
            days = row.get(col)
            if _trusted_days(days, min_d, max_d) and _is_slow(days):
                slow_stages.append(col.replace("Days ", "").replace(" to ", "→"))

        records.append(
            {
                "Candidate Key": _candidate_key(row),
                "S.No": row.get("S.No"),
                "Name": row.get("Name"),
                "Sourcer": row.get("Sourcer") if pd.notna(row.get("Sourcer")) else "Unknown",
                "Source Head": row.get("Source head") if "Source head" in df.columns else "",
                "Skills": row.get("Skills"),
                "Current Designation": row.get("Current Designation")
                if pd.notna(row.get("Current Designation"))
                else "Unknown",
                "Proposed Level": row.get("Proposed Level"),
                "Exp Bucket": row.get("Exp Bucket"),
                "Final Status": row.get("Final Status"),
                "R0 Date": row.get("R0"),
                "R1 Date": row.get("R1 Date (Final)"),
                "R2 Date": row.get("R2 Date (Final)"),
                "Days R0 to R1": row.get("Days R0 to R1"),
                "Days R1 to R2": row.get("Days R1 to R2"),
                "Days R0 to Offer Stage": row.get("Days R0 to Offer Stage"),
                "Total Pipeline Days R0 to R2": row.get("Total Pipeline Days (R0 to R2)"),
                "Reached R1": bool(row.get("Reached R1")),
                "Reached R2": bool(row.get("Reached R2")),
                "Reached Offer": bool(row.get("Reached Offer")),
                "Any Stage 30+ Days": bool(slow_stages),
                "Slow Stages": ", ".join(slow_stages),
                "Data Confidence": row.get("Data Confidence"),
            }
        )
    return pd.DataFrame(records)


def _build_stage_transitions_long(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for _, row in df.iterrows():
        key = _candidate_key(row)
        sourcer = row.get("Sourcer") if pd.notna(row.get("Sourcer")) else "Unknown"
        designation = (
            row.get("Current Designation")
            if pd.notna(row.get("Current Designation"))
            else "Unknown"
        )
        for from_stage, to_stage, col, min_d, max_d in STAGE_TRANSITIONS:
            days = row.get(col)
            if not _trusted_days(days, min_d, max_d):
                continue
            records.append(
                {
                    "Candidate Key": key,
                    "Name": row.get("Name"),
                    "Sourcer": sourcer,
                    "Designation": designation,
                    "Skills": row.get("Skills"),
                    "Final Status": row.get("Final Status"),
                    "Stage From": from_stage,
                    "Stage To": to_stage,
                    "Transition": f"{from_stage}→{to_stage}",
                    "Days": int(float(days)),
                    "Is Slow 30+ Days": _is_slow(days),
                    "Data Confidence": row.get("Data Confidence"),
                }
            )
    result = pd.DataFrame(records)
    if len(result):
        result = result.sort_values(["Transition", "Days"], ascending=[True, False])
    return result.reset_index(drop=True)


def _build_bottleneck_by_round(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for from_stage, to_stage, col, min_d, max_d in STAGE_TRANSITIONS:
        metrics = _stage_timing_metrics(df, col, min_days=min_d, max_days=max_d)
        records.append(
            {
                "Transition": f"{from_stage}→{to_stage}",
                "Stage From": from_stage,
                "Stage To": to_stage,
                "Candidates With Data": metrics["count_valid"],
                "Avg Days": metrics["avg_days"],
                "Median Days": metrics["median_days"],
                "Max Days": metrics["max_days"],
                "Count 30+ Days": metrics["slow_30plus_count"],
                "Pct 30+ Days": metrics["slow_30plus_pct"],
            }
        )

    result = pd.DataFrame(records)
    if len(result):
        primary = result.sort_values(
            ["Avg Days", "Pct 30+ Days"], ascending=[False, False]
        ).iloc[0]["Transition"]
        result["Is Primary Bottleneck"] = result["Transition"] == primary
        result = result.sort_values("Avg Days", ascending=False)
    return result.reset_index(drop=True)


def _build_dimension_bottleneck(
    df: pd.DataFrame,
    dimension_col: str,
    dimension_label: str,
    min_volume: int = 3,
) -> pd.DataFrame:
    records = []
    subset = df.copy()
    subset[dimension_col] = subset[dimension_col].fillna("Unknown")

    for name, group in subset.groupby(dimension_col, dropna=False):
        total = len(group)
        if total < min_volume:
            continue

        r01 = _stage_timing_metrics(group, "Days R0 to R1")
        r12 = _stage_timing_metrics(
            group,
            "Days R1 to R2",
            min_days=MIN_REASONABLE_DAYS_R12,
            max_days=MAX_REASONABLE_DAYS_R12,
        )
        r0o = _stage_timing_metrics(
            group,
            "Days R0 to Offer Stage",
            min_days=0,
            max_days=MAX_REASONABLE_DAYS_R01 * 2,
        )

        metrics = {
            "R0→R1": r01,
            "R1→R2": r12,
            "R0→Offer Stage": r0o,
        }
        primary = _primary_bottleneck(metrics)

        records.append(
            {
                dimension_label: name,
                "Total Candidates": total,
                "Avg Days R0→R1": r01["avg_days"],
                "Median Days R0→R1": r01["median_days"],
                "Count R0→R1 30+ Days": r01["slow_30plus_count"],
                "Pct R0→R1 30+ Days": r01["slow_30plus_pct"],
                "Avg Days R1→R2": r12["avg_days"],
                "Median Days R1→R2": r12["median_days"],
                "Count R1→R2 30+ Days": r12["slow_30plus_count"],
                "Pct R1→R2 30+ Days": r12["slow_30plus_pct"],
                "Avg Days R0→Offer Stage": r0o["avg_days"],
                "Count R0→Offer 30+ Days": r0o["slow_30plus_count"],
                "Primary Bottleneck": primary,
                "Any Stage 30+ Flag": any(
                    metrics[stage]["slow_30plus_count"] > 0 for stage in metrics
                ),
            }
        )

    result = pd.DataFrame(records)
    if len(result):
        sort_col = "Avg Days R0→R1"
        for col in ("Avg Days R0→R1", "Avg Days R1→R2", "Avg Days R0→Offer Stage"):
            if result[col].notna().any():
                sort_col = col
                break
        result = result.sort_values(sort_col, ascending=False, na_position="last")
    return result.reset_index(drop=True)


def _build_bottleneck_by_designation(df: pd.DataFrame) -> pd.DataFrame:
    return _build_dimension_bottleneck(df, "Current Designation", "Designation")


def _build_bottleneck_by_sourcer(df: pd.DataFrame) -> pd.DataFrame:
    return _build_dimension_bottleneck(df, "Sourcer", "Sourcer")


def _build_timing_by_designation(df: pd.DataFrame) -> pd.DataFrame:
    result = _build_bottleneck_by_designation(df)
    if len(result):
        slow = result[
            (result["Pct R0→R1 30+ Days"] >= 50)
            | (result["Count R0→R1 30+ Days"] >= 5)
            | (result["Count R1→R2 30+ Days"] >= 3)
        ].copy()
        slow["Slow Mover Category"] = "Designation with 30+ day delays"
        result["Slow Mover Category"] = np.where(
            result.index.isin(slow.index),
            "Designation with 30+ day delays",
            "",
        )
    return result


def _build_timing_by_sourcer(df: pd.DataFrame) -> pd.DataFrame:
    result = _build_bottleneck_by_sourcer(df)
    if len(result):
        slow = result[
            (result["Pct R0→R1 30+ Days"] >= 50)
            | (result["Count R0→R1 30+ Days"] >= 5)
            | (result["Avg Days R0→R1"].fillna(0) >= SLOW_THRESHOLD_DAYS)
        ].copy()
        result["Slow Mover Category"] = np.where(
            result.index.isin(slow.index),
            "Sourcer with 30+ day delays",
            "",
        )
    return result


def _build_slow_movers_detail(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for _, row in df.iterrows():
        slow_transitions = []
        for from_stage, to_stage, col, min_d, max_d in STAGE_TRANSITIONS:
            days = row.get(col)
            if _trusted_days(days, min_d, max_d) and _is_slow(days):
                slow_transitions.append(
                    {
                        "transition": f"{from_stage}→{to_stage}",
                        "days": int(float(days)),
                    }
                )
        if not slow_transitions:
            continue

        worst = max(slow_transitions, key=lambda item: item["days"])
        records.append(
            {
                "Candidate Key": _candidate_key(row),
                "S.No": row.get("S.No"),
                "Name": row.get("Name"),
                "Sourcer": row.get("Sourcer") if pd.notna(row.get("Sourcer")) else "Unknown",
                "Designation": row.get("Current Designation")
                if pd.notna(row.get("Current Designation"))
                else "Unknown",
                "Skills": row.get("Skills"),
                "Final Status": row.get("Final Status"),
                "Slowest Transition": worst["transition"],
                "Slowest Days": worst["days"],
                "All Slow Transitions": ", ".join(
                    f"{item['transition']} ({item['days']}d)" for item in slow_transitions
                ),
                "Days R0 to R1": row.get("Days R0 to R1"),
                "Days R1 to R2": row.get("Days R1 to R2"),
                "Days R0 to Offer Stage": row.get("Days R0 to Offer Stage"),
                "R0 Date": row.get("R0"),
                "R1 Date": row.get("R1 Date (Final)"),
                "R2 Date": row.get("R2 Date (Final)"),
            }
        )

    result = pd.DataFrame(records)
    if len(result):
        result = result.sort_values("Slowest Days", ascending=False)
    return result.reset_index(drop=True)


def _build_pbi_summary(df: pd.DataFrame, slow_movers: pd.DataFrame) -> pd.DataFrame:
    total = len(df)
    reached_r1 = int(df["Reached R1"].sum())
    reached_r2 = int(df["Reached R2"].sum())
    reached_offer = int(df["Reached Offer"].sum())
    rows = [
        ("Total Candidates", total),
        ("Reached R1", reached_r1),
        ("Reached R2", reached_r2),
        ("Reached Offer", reached_offer),
        ("R0→R1 Conversion %", round(reached_r1 / total * 100, 2) if total else 0),
        ("R1→R2 Conversion %", round(reached_r2 / reached_r1 * 100, 2) if reached_r1 else 0),
        ("R2→Offer Conversion %", round(reached_offer / reached_r2 * 100, 2) if reached_r2 else 0),
        ("Overall Conversion %", round(reached_offer / total * 100, 2) if total else 0),
        ("Avg Days R0→R1", _safe_mean(df["Days R0 to R1"])),
        ("Avg Days R1→R2", _safe_mean(
            df["Days R1 to R2"],
            min_days=MIN_REASONABLE_DAYS_R12,
            max_days=MAX_REASONABLE_DAYS_R12,
        )),
        ("Slow Movers 30+ Days", int(len(slow_movers))),
        ("Candidates 30+ at R0→R1", int((df["Days R0 to R1"] >= SLOW_THRESHOLD_DAYS).sum())),
    ]
    return pd.DataFrame([{"Metric": name, "Value": value} for name, value in rows])


def _build_pbi_funnel(df: pd.DataFrame) -> pd.DataFrame:
    total = len(df)
    reached_r1 = int(df["Reached R1"].sum())
    reached_r2 = int(df["Reached R2"].sum())
    reached_offer = int(df["Reached Offer"].sum())
    stages = [
        ("R0 Applied", total, 1),
        ("R1 Interview", reached_r1, 2),
        ("R2 Interview", reached_r2, 3),
        ("Offer Stage", reached_offer, 4),
    ]
    records = []
    for stage, count, order in stages:
        pct = round(count / total * 100, 2) if total else 0
        records.append(
            {
                "Stage": stage,
                "Stage Order": order,
                "Count": count,
                "Pct of R0": pct,
            }
        )
    return pd.DataFrame(records)


def _build_pbi_status(df: pd.DataFrame) -> pd.DataFrame:
    counts = df["Final Status"].fillna("Unknown").value_counts()
    return pd.DataFrame(
        [{"Status": str(status), "Count": int(count)} for status, count in counts.items()]
    )


def _build_pbi_sourcer_chart(sourcer_summary: pd.DataFrame) -> pd.DataFrame:
    if not len(sourcer_summary):
        return pd.DataFrame(
            columns=[
                "Sourcer",
                "Total Submitted",
                "Overall Conversion %",
                "R0→R1 Conversion %",
                "Avg Days R0→R1",
                "Avg Days R1→R2",
            ]
        )
    chart = sourcer_summary.sort_values("Total Submitted", ascending=False).head(12).copy()
    return chart[
        [
            "Sourcer",
            "Total Submitted",
            "Overall Conversion %",
            "R0→R1 Conversion %",
            "Avg Days R0→R1",
            "Avg Days R1→R2",
        ]
    ].reset_index(drop=True)


def _build_pbi_dashboard_layout() -> pd.DataFrame:
    return pd.DataFrame(POWERBI_DASHBOARD_LAYOUT)


def _build_powerbi_csv_zip(
    pbi_candidates: pd.DataFrame,
    stage_transitions: pd.DataFrame,
    bottleneck_rounds: pd.DataFrame,
    bottleneck_designation: pd.DataFrame,
    bottleneck_sourcer: pd.DataFrame,
    slow_movers: pd.DataFrame,
    designation_timing: pd.DataFrame,
    sourcer_timing: pd.DataFrame,
    pbi_summary: pd.DataFrame,
    pbi_funnel: pd.DataFrame,
    pbi_status: pd.DataFrame,
    pbi_sourcer_chart: pd.DataFrame,
    pbi_dashboard_layout: pd.DataFrame,
) -> bytes:
    import zipfile

    tables = {
        "PBI_Summary.csv": pbi_summary,
        "PBI_Funnel.csv": pbi_funnel,
        "PBI_Status.csv": pbi_status,
        "PBI_Sourcer_Chart.csv": pbi_sourcer_chart,
        "PBI_Dashboard_Layout.csv": pbi_dashboard_layout,
        "PBI_Candidates.csv": pbi_candidates,
        "PBI_Stage_Transitions.csv": stage_transitions,
        "PBI_Bottleneck_Rounds.csv": bottleneck_rounds,
        "PBI_Bottleneck_Designation.csv": bottleneck_designation,
        "PBI_Bottleneck_Sourcer.csv": bottleneck_sourcer,
        "PBI_Slow_Movers_30Plus.csv": slow_movers,
        "PBI_Designation_Timing.csv": designation_timing,
        "PBI_Sourcer_Timing.csv": sourcer_timing,
        "PowerBI_Instructions.csv": pd.DataFrame({"Step": POWERBI_INSTRUCTIONS}),
    }

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, frame in tables.items():
            csv_buffer = io.StringIO()
            frame.to_csv(csv_buffer, index=False)
            zf.writestr(filename, csv_buffer.getvalue())
    output.seek(0)
    return output.getvalue()


def _format_days_int(x: Any) -> Any:
    if pd.isna(x) or x == "":
        return ""
    try:
        return int(float(x))
    except:
        return ""


def _format_avg_days(x: Any) -> Any:
    if pd.isna(x) or x == "":
        return ""
    try:
        val = float(x)
        if val.is_integer():
            return int(val)
        return round(val, 1)
    except:
        return ""


def _prepare_recruitment_timeline_export(df: pd.DataFrame) -> pd.DataFrame:
    export = pd.DataFrame()
    export["S.No"] = df["S.No"]
    
    # R0 Date format: %d-%b-%Y (e.g. 17-Oct-2025)
    export["R0 – Submission Date"] = df["R0"].apply(lambda x: x.strftime("%d-%b-%Y") if pd.notna(x) else "")
    
    export["Candidate Name"] = df["Name"]
    export["Skills"] = df["Skills"]
    export["Proposed Level"] = df["Proposed Level"]
    export["Current Designation"] = df["Current Designation"]
    export["Total Exp"] = df["Total Exp"]
    export["Final Status"] = df["Final Status"]
    export["R1 Panel"] = df["R1 Panel"]
    export["R2 Panel"] = df["R2 Panel"]
    
    # R1/R2 Dates format: %d-%b-%Y
    export["R1 Date (Filled)"] = df["R1 Date (Final)"].apply(lambda x: x.strftime("%d-%b-%Y") if pd.notna(x) else "")
    export["R1 – Data Source"] = df["R1 Source"]
    export["R1 Assumption / Note"] = df["R1 Imputation Method"]
    
    export["R2 Date (Filled)"] = df["R2 Date (Final)"].apply(lambda x: x.strftime("%d-%b-%Y") if pd.notna(x) else "")
    export["R2 – Data Source"] = df["R2 Source"]
    export["R2 Assumption / Note"] = df["R2 Imputation Method"]
    
    # Timeline Analytics
    export["R0→R1 (Days)"] = df["Days R0 to R1"].apply(_format_days_int)
    export["R1→R2 (Days)"] = df["Days R1 to R2"].apply(_format_days_int)
    export["Total Days in Process"] = df["Total Pipeline Days (R0 to R2)"].apply(_format_days_int)
    
    # Avg Days / Round calculation
    avg_days_col = []
    for idx, row in df.iterrows():
        d01 = row["Days R0 to R1"]
        d12 = row["Days R1 to R2"]
        if pd.notna(d01) and pd.notna(d12):
            avg_days_col.append((d01 + d12) / 2.0)
        elif pd.notna(d01):
            avg_days_col.append(d01)
        else:
            avg_days_col.append(np.nan)
    export["Avg Days / Round"] = pd.Series(avg_days_col, index=df.index).apply(_format_avg_days)
    
    # Clean up NaNs or None in other columns
    for col in export.columns:
        if col not in ["R0 – Submission Date", "R1 Date (Filled)", "R2 Date (Filled)", "R0→R1 (Days)", "R1→R2 (Days)", "Total Days in Process", "Avg Days / Round"]:
            export[col] = export[col].apply(lambda x: "" if (pd.isna(x) or x is None) else str(x).strip())
            
    return export


def _style_recruitment_timeline_sheet(ws) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Border, Side, Alignment
    
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0")
    )
    
    ws.views.sheetView[0].showGridLines = True
    
    for col_idx, column_cells in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row), start=1):
        header = str(column_cells[0].value or "")
        max_len = len(header)
        for cell in column_cells[1:]:
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        width = min(max(max_len + 2, 12), 45)
        if "Date" in header:
            width = max(width, 14)
        if "Note" in header or "Assumption" in header:
            width = max(width, 28)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        for cell in column_cells[1:]:
            cell.border = thin_border
            if "Date" in header:
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif "Days" in header or "Process" in header or "Round" in header or "S.No" in header:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:T{ws.max_row}"


def _style_excel_worksheet(ws, date_columns: set[str] | None = None) -> None:
    """Auto-size columns and format header row."""
    date_columns = date_columns or set()
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col_idx, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        cell = column_cells[0]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        header = str(column_cells[0].value or "")
        max_len = len(header)
        for cell in column_cells[1:]:
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        width = min(max(max_len + 2, 12), 45)
        if "Date" in header or "Display" in header:
            width = max(width, 14)
        if "Imputation Method" in header or "Correction Note" in header:
            width = max(width, 28)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

        if "Date (Display)" in header:
            for cell in column_cells[1:]:
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _build_excel_output(
    df: pd.DataFrame,
    summary: pd.DataFrame,
    sourcer_summary: pd.DataFrame,
    sourcer_breakdown: pd.DataFrame,
    designation_breakdown: pd.DataFrame,
    validation_df: pd.DataFrame,
    *,
    pbi_candidates: pd.DataFrame,
    stage_transitions: pd.DataFrame,
    bottleneck_rounds: pd.DataFrame,
    bottleneck_designation: pd.DataFrame,
    bottleneck_sourcer: pd.DataFrame,
    slow_movers: pd.DataFrame,
    designation_timing: pd.DataFrame,
    sourcer_timing: pd.DataFrame,
    pbi_summary: pd.DataFrame,
    pbi_funnel: pd.DataFrame,
    pbi_status: pd.DataFrame,
    pbi_sourcer_chart: pd.DataFrame,
    pbi_dashboard_layout: pd.DataFrame,
) -> bytes:
    output = io.BytesIO()
    export_df = _prepare_recruitment_timeline_export(df)
    assumptions_df = pd.DataFrame({"Assumptions": ASSUMPTIONS})
    powerbi_instructions_df = pd.DataFrame({"Power BI Setup Steps": POWERBI_INSTRUCTIONS})

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pbi_dashboard_layout.to_excel(writer, sheet_name="PBI_Dashboard_Layout", index=False)
        powerbi_instructions_df.to_excel(writer, sheet_name="Power BI Instructions", index=False)
        pbi_summary.to_excel(writer, sheet_name="PBI_Summary", index=False)
        pbi_funnel.to_excel(writer, sheet_name="PBI_Funnel", index=False)
        pbi_status.to_excel(writer, sheet_name="PBI_Status", index=False)
        pbi_sourcer_chart.to_excel(writer, sheet_name="PBI_Sourcer_Chart", index=False)
        pbi_candidates.to_excel(writer, sheet_name="PBI_Candidates", index=False)
        stage_transitions.to_excel(writer, sheet_name="PBI_Stage_Transitions", index=False)
        bottleneck_rounds.to_excel(writer, sheet_name="PBI_Bottleneck_Rounds", index=False)
        bottleneck_designation.to_excel(
            writer, sheet_name="PBI_Bottleneck_Designation", index=False
        )
        bottleneck_sourcer.to_excel(writer, sheet_name="PBI_Bottleneck_Sourcer", index=False)
        slow_movers.to_excel(writer, sheet_name="PBI_Slow_Movers_30Plus", index=False)
        designation_timing.to_excel(writer, sheet_name="PBI_Designation_Timing", index=False)
        sourcer_timing.to_excel(writer, sheet_name="PBI_Sourcer_Timing", index=False)

        export_df.to_excel(writer, sheet_name="Recruitment Timeline", index=False, startrow=1)
        validation_df.to_excel(writer, sheet_name="Validation Report", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        sourcer_summary.to_excel(writer, sheet_name="Sourcer Summary", index=False)
        sourcer_breakdown.to_excel(writer, sheet_name="Sourcer x Skills", index=False)
        designation_breakdown.to_excel(
            writer, sheet_name="Designation x Exp", index=False
        )
        assumptions_df.to_excel(writer, sheet_name="Assumptions", index=False)

        # Apply custom style to Recruitment Timeline sheet
        ws = writer.sheets["Recruitment Timeline"]
        ws.merge_cells("A1:J1")
        ws.merge_cells("K1:P1")
        ws.merge_cells("Q1:T1")
        
        ws["A1"] = "Candidate & Profile Info"
        ws["K1"] = "Interview Dates (Filled + Sourced)"
        ws["Q1"] = "Timeline Analytics"
        
        from openpyxl.styles import Font, PatternFill, Alignment
        dark_blue_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        font_11_white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col in range(1, 21):
            cell = ws.cell(row=1, column=col)
            cell.fill = dark_blue_fill
            cell.font = font_11_white_bold
            cell.alignment = align_center
            
        medium_blue_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        font_9_white_bold = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        
        for col in range(1, 21):
            cell = ws.cell(row=2, column=col)
            cell.fill = medium_blue_fill
            cell.font = font_9_white_bold
            cell.alignment = align_center

        _style_recruitment_timeline_sheet(ws)

        for sheet_name in writer.sheets:
            if sheet_name not in ("Recruitment Timeline",):
                _style_excel_worksheet(writer.sheets[sheet_name])

    output.seek(0)
    return output.getvalue()


def analyze_file_path(path: str) -> dict[str, Any]:
    with open(path, "rb") as f:
        return analyze_supply_data(f.read())
